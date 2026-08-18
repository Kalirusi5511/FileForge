from docx import Document
from openpyxl import load_workbook


def docx_to_text(path):
    document = Document(path)

    return "\n".join(
        paragraph.text
        for paragraph in document.paragraphs
    )


def xlsx_to_text(path):
    workbook = load_workbook(path)

    output = []

    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            values = [
                str(cell)
                for cell in row
                if cell is not None
            ]

            if values:
                output.append(" ".join(values))

    return "\n".join(output)